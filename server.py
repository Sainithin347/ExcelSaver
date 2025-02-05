from flask import Flask, request, jsonify
import openpyxl
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

# The path to the Excel file
EXCEL_FILE = 'data.xlsx'

# Initialize the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Data'
    # Adding header row (Date, Place, Time, Kms)
    sheet.append(['Date', 'Place', 'Time', 'Kms'])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return open("D:\EVERYTHING\static\whatsapp.html").read()

@app.route('/save-to-excel', methods=['POST'])
def save_to_excel():
    # Get data from the request
    data = request.get_json()

    date = data['date']
    place = data['place']
    time = data['time']
    kms = data['kms']

    # Load the Excel file and select the active sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Append data to the next row
    sheet.append([date, place, time, kms])

    # Save the Excel file
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Data saved successfully!"})

if __name__ == '__main__':
    app.run(debug=True)
